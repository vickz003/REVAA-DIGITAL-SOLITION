from django.shortcuts import render, redirect,get_object_or_404, HttpResponse
from REVAA_RDS_APP.models import Leads
from django.utils import timezone
from django.contrib import messages
from datetime import datetime, timedelta
from django.db.models import Q
from django.shortcuts import redirect
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill
import pandas as pd 
from REVAA_RDS_APP.models import Leads
from django.contrib.auth import authenticate, login, logout
import json


# Create your views here.




def login_page(request):
    if request.user.is_authenticated:
        return redirect("index")
    if request.method == 'POST':
        username = request.POST["username"]
        password = request.POST["password"]
        user = authenticate(request,username=username, password=password)
        if user is not None:
            login(request,user)
            messages.success(request,"Successfully login....")
            return redirect("index")
    return render(request,"login.html")

def logout_page(request):
    logout(request)
    return redirect ("login")


def index(request):
    Dashboard = "active"
    today = datetime.now().date()
    tomorrow = today + timedelta(days=1)
    today_leads = Leads.objects.filter(Follow_up_date__range=(today, tomorrow))
    overall = len(today_leads)
    

    # today details
    todayleads = Leads.objects.filter(Follow_up_date__range=(today, today))
    today_leads_count = len(todayleads)
    todayleads_completed = len(Leads.objects.filter(
    Q(Follow_up_date=today) & Q(Status="Completed")))

    # Tomarrow
    Tomarrow = Leads.objects.filter(Follow_up_date__range=(tomorrow, tomorrow))
    Tomarrow_leads_count = len(Tomarrow)
    Tomarrowleads_completed = len(Leads.objects.filter(
    Q(Follow_up_date=tomorrow) & Q(Status="Completed")))

    # Over all
    overall_complited = todayleads_completed+Tomarrowleads_completed
    return render(request, 'index.html',
                  {'today_leads':today_leads,'overall':overall,'today_leads_count':today_leads_count,'todayleads_completed':todayleads_completed,
                 'Tomarrow_leads_count':Tomarrow_leads_count,'Tomarrowleads_completed':Tomarrowleads_completed,'overall_complited':overall_complited,
                 'Dashboard':Dashboard})

def leads(request):
    Leadss = "active"
    All_leads = Leads.objects.all()[::-1]
    last_lead = Leads.objects.last()
    service_list = str(last_lead.Services)
    services = [service.strip() for service in service_list.split(",")]
    leadheding = "Last Lead"
    container_hidden = 'container-hidden'

    return render(request, 'leads.html',{'services_list': services,'leads':All_leads,'last_lead':last_lead,'leadheding':leadheding,'Leadss':Leadss,'container_hidden':container_hidden})

def calls(request):
    return render(request, 'calls.html')

def lead_detail(request, lead_id):
    Leadss = "active"
    All_leads = Leads.objects.all()[::-1]
    leaddtails = Leads.objects.get(id = lead_id)
    leadheding = "Lead Details"
    hidden = "container-hidden"
    container_hidden = "container-hidden"

    lead = get_object_or_404(Leads, pk=lead_id)
    context = {'lead': lead,'leads':All_leads,'leaddtails':leaddtails,'leadheding':leadheding,'hidden':hidden,'Leadss':Leadss,'container_hidden':container_hidden}
    return render(request, 'leads.html', context)

def newleads(request):
    if request.method == "POST":
        name = request.POST["Name"]
        phone = request.POST["Contact_Number"]
        email = request.POST["Email_id"]
        location = request.POST["Location"]
        business_name = request.POST["Business_name"]
        budget = request.POST["Budget"]
        source = request.POST["Source"]
        selected_services_json = request.POST.get('selected_services')
        selected_services = json.loads(selected_services_json)
        services_text = ", ".join(selected_services)
        Source_Remarks = request.POST.get("Source_Remarks", "")
        status = request.POST["Status"]
        business_type = request.POST["Bussiness_types"]
        priority = request.POST["Priority"]
        current_date = timezone.now().date()
        followup_date = request.POST["Followupdate"]
        newlead = Leads(
            Name=name,
            Mobile_Number=phone,
            Email=email,
            Location=location,
            Business_Name=business_name,
            Budget=budget,
            Source=source,
            Services=services_text,
            Status=status,
            Business_Type=business_type,
            Priority=priority,
            Remark = Source_Remarks,
            Created_date = current_date,
            Follow_up_date=followup_date,
        )
        newlead.save()
        messages.success(request, 'Lead successfully added.')
        return redirect("leads")

    return render(request, 'leads.html')


# Download Leads

def exportdata(request):
    queryset = Leads.objects.all()
    wb = openpyxl.Workbook()
    ws = wb.active

    headers = ['Name', 'Mobile Number', 'Business Name', 'Business Type', 'Location', 'Budget', 'Follow-up Date', 'Priority', 'Status', 'Email', 'Source', 'Services', 'Remark', 'Created Date']

    header_style = Font(name='Arial', bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_style
        cell.fill = header_fill

    for row_num, lead in enumerate(queryset, start=2):
        data = [lead.Name, lead.Mobile_Number, lead.Business_Name, lead.Business_Type, lead.Location, lead.Budget, lead.Follow_up_date, lead.Priority, lead.Status, lead.Email, lead.Source, lead.Services, lead.Remark, lead.Created_date]
        for col_num, value in enumerate(data, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=value)

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=leads.xlsx'
    wb.save(response)

    return response

def editlead(request, lead_id):
    Leadss = "active"
    leadheding = "Lead Details"
    lead = Leads.objects.get(id=lead_id)
    service_list = str(lead.Services)
    services = [service.strip() for service in service_list.split(",")]

    return render(request, 'leads.html', {'lead': lead, 'Leadss': Leadss, 'leadheding': leadheding, 'services_list': services})


def updatedata(request, lead_id):
    mydata = Leads.objects.get(id=lead_id)
    if request.method == 'POST':
        name = request.POST["Name"]
        phone = request.POST["Contact_Number"]
        email = request.POST["Email_id"]
        location = request.POST["Location"]
        business_name = request.POST["Business_name"]
        budget = request.POST["Budget"]
        source = request.POST["Source"]
        services = request.POST.getlist('service')
        services_text = ",   ".join(services)
        Source_Remarks = request.POST.get("Source_Remarks", "")
        status = request.POST["Status"]
        business_type = request.POST["Bussiness_types"]
        priority = request.POST["Priority"]
        followup_date = request.POST["Followupdate"]

        mydata.Name = name
        mydata.Mobile_Number = phone
        mydata.Email = email
        mydata.Location = location
        mydata.Business_Name = business_name
        mydata.Budget = budget
        mydata.Source = source
        # mydata.Services = services_text
        mydata.Status = status
        mydata.Business_Type = business_type
        mydata.Priority = priority
        mydata.Follow_up_date = followup_date
        mydata.Remark = Source_Remarks
        mydata.save()
        messages.success(request, 'Lead successfully updated...')
        return redirect("index")
    
    return redirect('index')

def import_data(request):
    excel_file = request.FILES.get("ecxelfile")
    if excel_file:
        print("yes")
        data = pd.read_excel(excel_file)

        for index, row in data.iterrows():
            Leads.objects.create(
                Name = row['Name'], 
                Mobile_Number=row['ContactNo'], 
                Email=row['Mail Id'], 
                Business_Name=row['Business Name'], 
                Location=row['Location'], 
                Budget=row['Budget'], 
                Source=row['Source'],
                Services=row['Product'],  
                Status=row['Status'],
                Business_Type = row["Business Type"], 
                Priority = row['Priority'], 
                Remark = row['Remarks'], 
                Created_date = row['Date'], 
                Follow_up_date=row['EDD'],  
            )
        return redirect("newleads")
    return redirect("index")
